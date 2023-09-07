import openpyxl
workbooks = openpyxl.load_workbook('example.xlsx')
worksheet = workbooks.active
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value)
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active
    new_worksheet["A1"]="Name"
    new_worksheet["B1"]="Age"
    new_worksheet["C1"]="city"
    new_worksheet["D1"]="Country"
    new_worksheet.save("example.xlsx")